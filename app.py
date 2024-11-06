from tkinter import * 
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
from fpdf import FPDF
from tkinter import filedialog
import pandas as pd

#global proximo_ID = 1

def verifica_se_existe_cadastro(doc):
    #ler a tabela do excel
    df = pd.read_excel('funcionarios.xlsx')

    #verifica se o documento já existe
    if doc in df['Documento'].values:
        print("este cadastro ja existe.")
        return True
    else:
        print("cadastro disponivel.")
        return False
    
#Essa funcao seleciona 1 item da treeview e apaga
def apagar(treeview):  
    # Obtém a seleção atual
    selecionado = treeview.selection()
    
    if not selecionado:
        # Se não houver seleção, exibe uma mensagem
        print("Nao tem item selecionado")
        messagebox.showinfo("Nenhum item selecionado", "Por favor, selecione um item para deletar.")
        return

    # Deleta cada item selecionado
    for item in selecionado:
        treeview.delete(item)
    print("Item deletado", "O item selecionado foi deletado com sucesso.")
    messagebox.showinfo("Item deletado", "O item selecionado foi deletado com sucesso.")


def registrar():
    cadastro = inserir_cadastro.get().strip()
    destino = inserir_destino.get().strip()
    obs = inserir_obs.get().strip()
    
    # Verifica se os campos obrigatórios estão preenchidos
    if not cadastro:
        messagebox.showwarning("Campo Vazio", "Por favor, preencha o campo 'Cadastro'.")
        return
    if not destino:
        messagebox.showwarning("Campo Vazio", "Por favor, preencha o campo 'Destino'.")
        return

    # Lê a tabela do Excel
    df = pd.read_excel('funcionarios.xlsx')

    # Verifica se o cadastro existe
    if cadastro not in df['Documento'].values:
        messagebox.showerror("Cadastro Não Encontrado", "Esse cadastro não existe.")
        return
    
    # Obtém os dados do cadastro
    linha = df.loc[df['Documento'] == cadastro].iloc[0]
    # Obtém a hora atual
    hora_atual = datetime.now().strftime("%H:%M:%S")
    if obs == "OBS":
        obs = ""
    linha = list(linha)

    linha.append(destino)
    linha.append(hora_atual)
    linha.append(obs)
    

        
    print(list(linha))

    #imprime a lista na treewiew
    treeview.insert("","end", values=linha)
    # Obtenha outros dados conforme necessário
    
    


    # Limpa os campos
    inserir_cadastro.delete(0, tk.END)
    inserir_destino.delete(0, tk.END)
    inserir_obs.delete(0, tk.END)
    


    


def salvar_pdf():
    # Abre uma janela de diálogo para o usuário escolher o local e nome do arquivo
    filepath = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        title="Salvar como"
    )
    
    if not filepath:  # Se o usuário cancelar o diálogo, sai da função
        return
    
    # Cria o PDF
    pdf = FPDF(orientation="L")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Arial", size=10)

    # Adiciona o cabeçalho da tabela
    colunas = ["Documento", "Nome", "Validade", "Militar", "Veículo", "Placa", "Cor", "Destino", "Hora", "OBS"]
    pdf.set_fill_color(200, 200, 200)  # Cor de fundo para o cabeçalho
    pdf.set_text_color(0, 0, 0)  # Cor do texto

    # Define a largura das colunas no PDF
    largura_colunas = [25, 40, 20, 10, 30, 20, 15, 20, 20, 30]
    for col, largura in zip(colunas, largura_colunas):
        pdf.cell(w=largura, h=10, txt=col, border=1, align='C', fill=True)
    pdf.ln()  # Nova linha

    # Adiciona os dados da Treeview ao PDF
    for row_id in treeview.get_children():
        row = treeview.item(row_id)["values"]
        for item, largura in zip(row, largura_colunas):
            pdf.cell(w=largura, h=10, txt=str(item), border=1, align='C')
        pdf.ln()  # Nova linha para cada registro

    # Salva o PDF no caminho especificado
    pdf.output(filepath)
    print(f"PDF salvo em: {filepath}")



#Backend
def carregar_dados():
    path = r"C:\Users\GG\Desktop\app cadastro de entrada\funcionarios.xlsx"
    arquivo_tabela = openpyxl.load_workbook(path)
    tabela = arquivo_tabela.active

    lista_de_valores = list(tabela.values)
    print(lista_de_valores)

    for nome_na_tabela in colunas:
        treeview.heading(nome_na_tabela, text=nome_na_tabela)

    #for valor_da_tabela in lista_de_valores[0:]:
     #   treeview.insert("",tk.END,values=valor_da_tabela)



def cadastrar(documento, nome, validade, militar_var, veiculo, placa, cor_entrada, nova_janela):

    doc = documento.get()
    nm = nome.get()
    val = validade.get()
    mil = militar_var.get()
    vec = veiculo.get()
    pl = placa.get()
    cor = cor_entrada.get()
    verifica_se_existe_cadastro(doc)
    if verifica_se_existe_cadastro(doc) == True:
        print("Esse cadastro ja esxite")
        nova_janela.destroy()
    else:
        print("Cadastro Realizado:")
        print(f"Documento: {doc}")
        print(f"Nome: {nm}")
        print(f"Validade: {val}")
        print(f"É Militar: {mil}")
        print(f"Veículo: {vec}")
        print(f"Placa: {pl}")
        print(f"Cor: {cor}")
        

        #salvar dados na folha do excel
        path = r"C:\Users\GG\Desktop\app cadastro de entrada\funcionarios.xlsx"
        arquivo_tabela = openpyxl.load_workbook(path)
        tabela = arquivo_tabela.active
        valor_das_linhas = [doc,nm,val,mil,vec,pl,cor]
        tabela.append(valor_das_linhas)
        arquivo_tabela.save(path)

        #inserir dados salvo na treeview
        #treeview.insert("",tk.END,values=valor_das_linhas)
        
       


        print("cadastro realizado com sucesso")
        inserir_cadastro.insert(0,doc)



        #limpar e colocar os dados iniciais dos nossos campos
        nova_janela.destroy()

#---------------------------------------------------------------------------------------------

def mudar_tema():
    if botao_tema.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")


def formatar_placa(event,placa):

    # Obtém o valor da placa até agora
    texto = placa.get()

    # Remove caracteres não alfabéticos e não numéricos
    texto = ''.join(filter(lambda x: x.isalnum(), texto))

    # Condição para as três primeiras posições serem letras
    if len(texto) >= 1 and not texto[0].isalpha():
        texto = texto[1:]  # Apaga o primeiro caractere se não for letra
    if len(texto) >= 2 and not texto[1].isalpha():
        texto = texto[:1] + texto[2:]  # Apaga o segundo caractere se não for letra
    if len(texto) >= 3 and not texto[2].isalpha():
        texto = texto[:2] + texto[3:]  # Apaga o terceiro caractere se não for letra

    # Condição para a quarta posição ser um número
    if len(texto) >= 4 and not texto[3].isdigit():
        texto = texto[:3]  # Apaga o quarto caractere se não for número

    # Condição para a quinta posição (letra ou número é permitido, então não há validação aqui)

    # Condição para a sexta e sétima posições serem números
    if len(texto) >= 6 and not texto[5].isdigit():
        texto = texto[:5]  # Apaga o sexto caractere se não for número
    if len(texto) >= 7 and not texto[6].isdigit():
        texto = texto[:6]  # Apaga o sétimo caractere se não for número

    # Limita o texto a no máximo 7 caracteres
    texto = texto[:7]

    # Atualiza o campo com o texto formatado
    placa.delete(0, tk.END)
    placa.insert(0, texto)
    
def formatar_documento(event,documento):
    # Obtém o texto digitado até agora
    texto = documento.get()
    # Remove qualquer caractere que não seja dígito para simplificar
    texto = ''.join(filter(str.isdigit, texto))
     # Remove zeros à esquerda
    texto = texto.lstrip('0')  # ou texto = str(int(texto)) se preferir
    # Atualiza o campo de entrada com o texto formatado
    documento.delete(0, tk.END)
    documento.insert(0, texto)
    #documento.insert(texto)


def formatar_validade(event,validade):
    
    # Obtém o texto digitado até agora
    texto = validade.get()

    # Remove qualquer caractere que não seja dígito para simplificar
    texto = ''.join(filter(str.isdigit, texto))

    # Formata automaticamente no estilo DD/MM/AAAA
    if len(texto) > 2 and len(texto) <= 4:
        texto = texto[:2] + '/' + texto[2:]
    elif len(texto) > 4:
        texto = texto[:2] + '/' + texto[2:4] + '/' + texto[4:8]

    # Validação do mês e do dia
    if len(texto) >= 5:
        dia = int(texto[:2])
        mes = int(texto[3:5]) if len(texto) > 2 else 0  # O mês começa no índice 3

        # Verifica se o mês é maior que 12
        if mes > 12:
            mes = 12
            texto = f"{dia:02}/{mes:02}/{texto[6:]}"  # Corrige o mês

        # Verifica se o dia é maior que 31
        if dia > 31:
            dia = 31
            texto = f"{dia:02}/{mes:02}/{texto[6:]}"  # Corrige o dia

    # Atualiza o campo com o texto formatado
    validade.delete(0, tk.END)
    validade.insert(0, texto)

def nova_janela_cadastro():
    # Cria uma nova janela (janela secundária)
    nova_janela = tk.Toplevel(root)
    nova_janela.title("Cadastro")  # Título da janela (sem exibição na interface principal)

    # Frame principal com texto de instrução
    frame_principal = ttk.Frame(nova_janela)
    frame_principal.pack(padx=10, pady=10, fill="both")

    # Texto de instrução no topo
    label_instrucao = ttk.Label(frame_principal, text="Insira os Dados", font=("Arial", 14))
    label_instrucao.grid(row=0, column=0, columnspan=2, pady=(0, 10), sticky="ew")
    label_instrucao.config(anchor="center") # Centraliza o texto dentro do Label

    # Documento
    ttk.Label(frame_principal, text="Documento: ").grid(row=1, column=0, padx=(5), pady=5, sticky="w")
    documento = ttk.Entry(frame_principal)
    documento.insert(0, "Documento")
    documento.bind("<FocusIn>", lambda e: documento.delete(0, "end"))
    documento.bind("<KeyRelease>", lambda event: formatar_documento(event, documento))
    documento.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    

    # Nome
    ttk.Label(frame_principal, text="Nome: ").grid(row=2, column=0, padx=(5), pady=5, sticky="w")
    nome = ttk.Entry(frame_principal)
    nome.insert(0, "Nome")
    nome.bind("<FocusIn>", lambda e: nome.delete(0, "end"))
    nome.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
    

    # Validade do Documento
    ttk.Label(frame_principal, text="Validade do Documento:").grid(row=3, column=0, padx=(5), pady=5, sticky="w")
    validade = ttk.Entry(frame_principal)
    validade.insert(0, "Validade")
    validade.bind("<FocusIn>", lambda e: validade.delete(0, "end"))
    validade.bind("<KeyRelease>", lambda event: formatar_validade(event, validade))
    validade.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
    
    

    # É Militar (Sim ou Não)
    ttk.Label(frame_principal, text="Militar?").grid(row=4, column=0, padx=(5), pady=5, sticky="w")
    militar_var = tk.StringVar(value="N")  # Valor padrão "Não"
    radio_sim = ttk.Radiobutton(frame_principal, text="Sim", variable=militar_var, value="S")
    radio_nao = ttk.Radiobutton(frame_principal, text="Não", variable=militar_var, value="N")
    radio_sim.grid(row=4, column=1, padx=(5, 25), pady=5, sticky="e")
    radio_nao.grid(row=4, column=1, padx=(25, 0), pady=5, sticky="w")  # Alinhamento dos botões
    

    # Veículo
    ttk.Label(frame_principal, text="Veículo: ").grid(row=5, column=0, padx=(5), pady=5, sticky="w")
    veiculo = ttk.Entry(frame_principal)
    veiculo.insert(0, "Veículo")
    veiculo.bind("<FocusIn>", lambda e: veiculo.delete(0, "end"))
    veiculo.grid(row=5, column=1, padx=5, pady=5, sticky="ew")
    

    # Placa
    ttk.Label(frame_principal, text="Placa: ").grid(row=6, column=0, padx=(5), pady=5, sticky="w")
    placa = ttk.Entry(frame_principal)
    placa.insert(0, "Placa")
    placa.bind("<FocusIn>", lambda e: placa.delete(0, "end"))
    placa.bind("<KeyRelease>", lambda event: formatar_placa(event, placa))
    placa.grid(row=6, column=1, padx=5, pady=5, sticky="ew")
    
    

    # Cor
    ttk.Label(frame_principal, text="Cor: ").grid(row=7, column=0, padx=(5), pady=5, sticky="w")
    cor_entrada = ttk.Combobox(frame_principal, values=["Branco", "Preto", "Cinza", "Prata", "Vermelho", "Verde", "Amarelo", "Azul", "Dourado", "Vinho"])
    cor_entrada.grid(row=7, column=1, padx=5, pady=5, sticky="ew")
    

    # Botão para salvar o cadastro
    botao_salvar = ttk.Button(frame_principal, text="Salvar Cadastro", command=lambda: cadastrar(documento, nome, validade, militar_var, veiculo, placa, cor_entrada, nova_janela))
    botao_salvar.grid(row=8, column=0, columnspan=2, padx=5, pady=(10, 5), sticky="ew")

    # Função para fechar a janela
    def fechar_janela(event=None):
        nova_janela.destroy()

    nova_janela.bind("<Escape>", fechar_janela)

    
    
    



#---------------------------------------------------------------------------------


#iniciando nossa janela

root = tk.Tk()
#titulo do app
root.title("Cadastro de Entrada")

#configuração do nosso tema
style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")

style.theme_use("forest-dark")

#criando o espacos para os objetos frame principal
frame_principal = ttk.Frame( root)
frame_principal.pack()

widgets_frame = ttk.LabelFrame( frame_principal, text="Insira os dados")
widgets_frame.grid(row=0, column=0, padx=10, pady=10)

#inserir cadastro
label_i_c = ttk.Label( widgets_frame, text="Documento")
label_i_c.grid(row=0, column=0, padx=5, pady=0,sticky="w")
inserir_cadastro = ttk.Entry(widgets_frame)
inserir_cadastro.insert(0,"")
inserir_cadastro.bind("<FocusIn>", lambda e: inserir_cadastro.delete(0, "end") if inserir_cadastro.get() == "Cadastro" else None)
inserir_cadastro.grid(row=1,column=0,padx=5,pady=5,sticky="ew")
inserir_cadastro.bind("<KeyRelease>", lambda event: formatar_documento(event, inserir_cadastro))

#Destino
label_i_d = ttk.Label( widgets_frame, text="Destino:")
label_i_d.grid(row=2, column=0, padx=5, pady=0,sticky="w")
inserir_destino = ttk.Entry(widgets_frame)
inserir_destino.insert(0,"")
inserir_destino.bind("<FocusIn>", lambda e:inserir_destino.delete("0","end"))
inserir_destino.grid(row=3,column=0,padx=5,pady=5,sticky="ew")

#Obs
label_i_o = ttk.Label( widgets_frame, text="OBS:")
label_i_o.grid(row=4, column=0, padx=5, pady=0,sticky="w")
inserir_obs = ttk.Entry(widgets_frame)
inserir_obs.insert(0,"")
inserir_obs.bind("<FocusIn>", lambda e:inserir_obs.delete("0","end"))
inserir_obs.grid(row=5,column=0,padx=5,pady=(5,50),sticky="ew")

#botao para registro
botao_registro = tk.Button(widgets_frame, text="Registro", command=registrar)
botao_registro.grid(row=6,column=0,padx=5,pady=5,sticky="ew")

#botao para cadastro
botao_novo_cadastro = tk.Button(widgets_frame, text="Novo Cadastro", command=nova_janela_cadastro)
botao_novo_cadastro.grid(row=7,column=0,padx=5,pady=5,sticky="ew")

#botao para apagar registro
botao_apagar_registro = tk.Button(widgets_frame, text="Apagar", command=lambda: apagar (treeview))
botao_apagar_registro.grid(row=8,column=0,padx=5,pady=5,sticky="ew")

#separador
separador = ttk.Separator(widgets_frame)
separador.grid(row=9,column=0,padx=5,pady=5,sticky="ew")

#botao salvar
botao_salvar = tk.Button(widgets_frame, text="Salvar",command=salvar_pdf)
botao_salvar.grid(row=10,column=0,padx=5,pady=5,sticky="ew")



#switch do tema
botao_tema = ttk.Checkbutton(widgets_frame,text="Tema",style="Switch",command=mudar_tema)
botao_tema.grid(row=11,column=0,padx=50,pady=(5,10),sticky="ew")





#criando a treeview do banco de dados
treeFrame = ttk.Frame(frame_principal)
treeFrame.grid(row=0, column=1,padx=(0,20),pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right",fill="y")

colunas = ("Documento","Nome","Validade","Militar","Veiculo","Placa","Cor","Destino","Hora","OBS")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=colunas, height=20)


treeview.column("Documento",width=100,anchor="center")
treeview.column("Nome",width=150,anchor="center")
treeview.column("Validade",width=60,anchor="center")
treeview.column("Militar",width=35,anchor="center")
treeview.column("Veiculo",width=60,anchor="center")
treeview.column("Placa",width=70,anchor="center")
treeview.column("Cor",width=100,anchor="center")
treeview.column("Destino",width=100,anchor="center")
treeview.column("Hora",width=53,anchor="center")
treeview.column("OBS",width=100,anchor="center")



treeview.pack()
treeScroll.config(command=treeview.yview)


carregar_dados()


#rodando nossa janela
root.mainloop()
"""
#na segunda tela teremos o cadastro dos seguintes elementos


"""
#validade

#émilitar?
#veiculo
#cor
#placa
#destino
#hora





#botao editar cadastro
#
#
#
#


#
#
#
#
#

